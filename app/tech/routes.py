import os
import ldap
from flask import render_template, request, jsonify, current_app
from openpyxl import load_workbook
from . import tech

# ==============================================================================
# 1. CONFIGURATION DES CHEMINS (Version Robuste)
# ==============================================================================
# Récupère le dossier où se trouve ce fichier (app/tech/)
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
# Remonte de deux niveaux pour trouver la racine du projet (assane18-ilvmintra/)
PROJECT_ROOT = os.path.dirname(os.path.dirname(CURRENT_DIR))
# Définit le dossier data et le fichier stock
DATA_DIR = os.path.join(PROJECT_ROOT, 'data')
DEFAULT_INVENTORY = os.path.join(DATA_DIR, 'stock.xlsx')

print(f"--- DEBUG: Chemin du fichier Excel visé : {DEFAULT_INVENTORY}")

# ==============================================================================
# 2. CONFIGURATION LDAP (En dur pour être sûr)
# ==============================================================================
LDAP_SERVER = 'ldap://192.168.1.9'
LDAP_BASE_DN = 'DC=ilvm,DC=lan'
LDAP_USER = 'CN=Admin Intra,CN=Users,DC=ilvm,DC=lan'
LDAP_PASSWORD = 'gq!nsXPYsM!LmFh4'

# ==============================================================================
# 3. ROUTES
# ==============================================================================

@tech.route('/generateur')
def home():
    return render_template('tech/generateur_pret.html')

def normalize_header(header):
    """Nettoie les entêtes du fichier Excel"""
    h = str(header).upper().strip()
    if 'SERIE' in h or 'SERIAL' in h or h == 'SN' or 'S/N' in h: return 'SN'
    if 'TYPE' in h: return 'Type'
    if 'MARQUE' in h: return 'Marque'
    if 'MODEL' in h: return 'Modele'
    if 'IMEI' in h: return 'IMEI'
    if 'PIN' in h: return 'PIN'
    return h 

@tech.route('/upload_inventory', methods=['POST'])
def upload_inventory():
    print("--- DEBUG: Tentative d'upload fichier")
    if 'file' not in request.files: 
        return jsonify({"success": False, "message": "Aucun fichier reçu"})
    
    file = request.files['file']
    if not file.filename.endswith('.xlsx'): 
        return jsonify({"success": False, "message": "Format invalide (.xlsx requis)"})
    
    try:
        # On s'assure que le dossier data existe, sinon on le crée
        if not os.path.exists(DATA_DIR):
            os.makedirs(DATA_DIR)
            print(f"--- DEBUG: Dossier {DATA_DIR} créé.")

        file.save(DEFAULT_INVENTORY)
        print("--- DEBUG: Fichier sauvegardé avec succès.")
        return jsonify({"success": True, "message": "Base mise à jour sur le serveur !"})
    except Exception as e:
        print(f"--- ERROR Upload: {str(e)}")
        return jsonify({"success": False, "message": str(e)})

@tech.route('/get_inventory')
def get_inventory():
    print(f"--- DEBUG: Lecture inventaire depuis {DEFAULT_INVENTORY}")
    
    if not os.path.exists(DEFAULT_INVENTORY):
        print("--- ERROR: Fichier introuvable sur le disque.")
        return jsonify([])

    try:
        wb = load_workbook(DEFAULT_INVENTORY, data_only=True)
        ws = wb.active
        data = []
        
        # Récupération des entêtes
        headers = [normalize_header(cell.value) if cell.value else "Inconnu" for cell in ws[1]]
            
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = {}
            for i, cell_value in enumerate(row):
                if i < len(headers):
                    val = str(cell_value).strip() if cell_value is not None else ""
                    # Nettoyage des '.0' pour les nombres convertis en texte
                    if headers[i] in ['IMEI', 'PIN', 'SN'] and val.endswith('.0'):
                        val = val[:-2]
                    item[headers[i]] = val
            
            # Logique spécifique : si Modèle vide mais Marque présente
            if 'Marque' in item and ('Modele' not in item or not item['Modele']):
                item['Modele'] = item['Marque']

            # On garde seulement les lignes avec un identifiant
            has_id = ('SN' in item and item['SN']) or ('IMEI' in item and item['IMEI'])
            if has_id:
                data.append(item)
        
        print(f"--- DEBUG: {len(data)} articles chargés.")
        return jsonify(data)

    except Exception as e:
        print(f"--- ERROR Excel Read: {str(e)}")
        return jsonify({"error": str(e)})

@tech.route('/get_user/<username>')
def get_user(username):
    print(f"--- DEBUG: Recherche LDAP pour {username}")
    try:
        # Initialisation
        l = ldap.initialize(LDAP_SERVER)
        l.protocol_version = ldap.VERSION3
        l.set_option(ldap.OPT_REFERRALS, 0)
        
        # Connexion (Bind)
        try:
            l.simple_bind_s(LDAP_USER, LDAP_PASSWORD)
            print("--- DEBUG: Connexion LDAP (Bind) réussie.")
        except ldap.INVALID_CREDENTIALS:
            print("--- ERROR: Mot de passe Admin LDAP incorrect.")
            return jsonify({"success": False, "message": "Erreur Auth LDAP (Mot de passe Admin)"})
        except Exception as e:
            print(f"--- ERROR LDAP Bind: {str(e)}")
            return jsonify({"success": False, "message": f"Erreur Connexion Serveur: {str(e)}"})

        # Recherche
        search_filter = f"(sAMAccountName={username})"
        attributes = ['displayName', 'mail', 'department', 'telephoneNumber']
        
        result = l.search_s(LDAP_BASE_DN, ldap.SCOPE_SUBTREE, search_filter, attributes)

        if result and len(result) > 0 and result[0][1]:
            user_data = result[0][1]
            # Décodage sécurisé
            def decode_attr(attr_name):
                val = user_data.get(attr_name, [b''])
                if val:
                    return val[0].decode('utf-8', errors='ignore')
                return ""

            response_data = {
                "success": True,
                "nom": decode_attr('displayName'),
                "mail": decode_attr('mail'),
                "service": decode_attr('department'),
                "telephone": decode_attr('telephoneNumber')
            }
            print(f"--- DEBUG: Utilisateur trouvé: {response_data['nom']}")
            return jsonify(response_data)
        else:
            print("--- DEBUG: Utilisateur introuvable.")
            return jsonify({"success": False, "message": "Utilisateur introuvable"})
            
    except Exception as e:
        print(f"--- ERROR LDAP Search: {str(e)}")
        return jsonify({"success": False, "message": str(e)})
